"""
Motor de evaluación RIPS vs Resolución 3280/2018
Procesa JSON de RIPS y valida contra el estándar parametrizado
"""
import json, datetime, re
from pathlib import Path
from collections import defaultdict

CONFIG_PATH = Path(__file__).parent / "config" / "res3280_cups.json"

def _load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)

def _calcular_edad(fecha_nac_str, fecha_ref=None):
    """Calcula edad en años desde fecha de nacimiento (string YYYY-MM-DD o DD/MM/YYYY)"""
    if not fecha_nac_str:
        return None
    try:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
            try:
                fn = datetime.datetime.strptime(str(fecha_nac_str)[:10], fmt).date()
                break
            except ValueError:
                continue
        else:
            return None
        ref = fecha_ref or datetime.date.today()
        edad = ref.year - fn.year - ((ref.month, ref.day) < (fn.month, fn.day))
        return edad
    except Exception:
        return None

def _grupo_edad(edad, cursos_de_vida):
    """Retorna el ID del grupo de vida según la edad"""
    if edad is None:
        return None
    for cv in cursos_de_vida:
        if cv["edad_min"] <= edad <= cv["edad_max"]:
            return cv["id"]
    return None

def _norm(s):
    """Normaliza string: mayúsculas, sin tildes"""
    if not s:
        return ""
    s = str(s).upper().strip()
    for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N")]:
        s = s.replace(a, b)
    return s

def _paciente_key(registro, fm):
    """Llave única de paciente: tipo_doc + num_doc"""
    return (str(registro.get(fm["tipo_doc"],"") or "").strip(),
            str(registro.get(fm["num_doc"],"") or "").strip())

class RIPSEvaluator:
    def __init__(self, periodo_ref=None):
        self.cfg = _load_config()
        self.periodo_ref = periodo_ref or datetime.date.today()
        self._usuarios = {}      # clave: (tipo_doc, num_doc) → {edad, grupo}
        self._archivos = {}      # nombre → lista de dicts
        self._raw_usuarios = []  # datos crudos hasta que se calculen los grupos

    # ── Carga de datos ─────────────────────────────────────────────────────────
    def cargar_usuarios(self, data: list):
        """Guarda los usuarios crudos; los grupos se calculan después de cargar todos los archivos."""
        self._raw_usuarios = data

    def _calcular_grupos(self):
        """
        Calcula el grupo de vida de cada paciente usando la fecha más reciente
        en que aparece en CUALQUIER archivo RIPS — igual que la fórmula Excel:
          f = MAX(MAX.SI.CONJUNTO(consultas!E, consultas!C, numDoc), ...)
          edad = SIFECHA(fechaNac, f, "Y")
          grupo = SI.CONJUNTO(edad<=5,"PRIMERA INFANCIA", ...)
        """
        fm  = self.cfg["field_mappings"]["usuarios"]
        cvs = self.cfg["cursos_de_vida"]

        # Campos de fecha y numDoc por archivo
        FECHA_FIELD = {
            "consultas":      "fechaInicioAtencion",
            "procedimientos": "fechaInicioAtencion",
            "medicamentos":   "fechaDispensacionMedicamento",
            "otrosServicios": "fechaInicioAtencion",
        }

        # Paso 1: max_fecha por numDoc (sin tipo_doc, igual que Excel columna C)
        max_fecha_by_num: dict = {}
        for archivo, registros in self._archivos.items():
            fecha_key = FECHA_FIELD.get(archivo, "fechaInicioAtencion")
            for r in registros:
                num_doc = str(r.get("numDocumentoIdentificacion", "") or "").strip()
                if not num_doc:
                    continue
                fecha_str = str(r.get(fecha_key, "") or "")[:10]
                try:
                    fecha = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
                    if num_doc not in max_fecha_by_num or fecha > max_fecha_by_num[num_doc]:
                        max_fecha_by_num[num_doc] = fecha
                except Exception:
                    pass

        # Paso 2: asignar grupo usando la fecha real de atención
        self._usuarios = {}
        for u in self._raw_usuarios:
            k = _paciente_key(u, fm)
            fn_str  = u.get(fm["fecha_nacimiento"])
            num_doc = str(u.get(fm["num_doc"], "") or "").strip()
            # Fecha de referencia = última atención encontrada o periodo_ref como fallback
            ref_date = max_fecha_by_num.get(num_doc, self.periodo_ref)
            edad  = _calcular_edad(fn_str, ref_date)
            grupo = _grupo_edad(edad, cvs)
            self._usuarios[k] = {
                "edad":  edad,
                "grupo": grupo,
                "sexo":  str(u.get("codSexo", "")).strip().upper()
            }

    def cargar_archivo(self, nombre: str, data: list):
        """Carga un tipo de archivo RIPS (consultas, procedimientos, etc.)"""
        nombre = nombre.lower().replace(".json","")
        # Normalizar nombre: aceptar variantes
        for alias, canon in [
            ("consultaambulatorio","consultas"),
            ("consultaambulatorias","consultas"),
            ("procedimiento","procedimientos"),
            ("otroservicio","otrosServicios"),
            ("otrosservicio","otrosServicios"),
            ("medicamento","medicamentos"),
            ("usuario","usuarios"),
        ]:
            if alias in nombre:
                nombre = canon; break
        if nombre == "usuarios":
            self.cargar_usuarios(data)
        else:
            self._archivos[nombre] = data

    # ── Enriquecimiento: agrega grupo de vida a cada registro ─────────────────
    def _enriquecer(self, registros: list, archivo_nombre: str) -> list:
        fm_users = self.cfg["field_mappings"]["usuarios"]
        fm_arch  = self.cfg["field_mappings"].get(archivo_nombre, {})
        tipo_doc_key = fm_arch.get("tipo_doc", "tipoDocumentoIdentificacion")
        num_doc_key  = fm_arch.get("num_doc",  "numDocumentoIdentificacion")
        enriched = []
        for r in registros:
            k = (str(r.get(tipo_doc_key,"") or "").strip(),
                 str(r.get(num_doc_key, "") or "").strip())
            info = self._usuarios.get(k, {"edad": None, "grupo": None, "sexo": ""})
            rec = dict(r)
            rec["__grupo"] = info["grupo"]
            rec["__edad"]  = info["edad"]
            rec["__sexo"]  = info["sexo"]
            enriched.append(rec)
        return enriched

    # ── Conteo de una actividad para un grupo de vida ─────────────────────────
    def _contar_actividad(self, actividad_cfg: dict, grupo_id: str) -> int:
        archivo  = actividad_cfg["archivo"]
        cups_set = set(str(c).strip() for c in actividad_cfg.get("cups", []))
        finalidades = set(str(f).strip() for f in actividad_cfg.get("finalidad", []))
        busqueda_nombre = [_norm(n) for n in actividad_cfg.get("busqueda_nombre", [])]
        deduplicar = actividad_cfg.get("deduplicar", True)

        registros_raw = self._archivos.get(archivo, [])
        if not registros_raw:
            return 0

        registros = self._enriquecer(registros_raw, archivo)

        contados = set() if deduplicar else []
        fm = self.cfg["field_mappings"].get(archivo, {})

        cups_field      = fm.get("cups", "codProcedimiento")
        finalidad_field = fm.get("finalidad", "finalidadTecnologiaSalud")
        nombre_field    = fm.get("nombre", "nomTecnologiaSalud")
        tipo_doc_key    = fm.get("tipo_doc", "tipoDocumentoIdentificacion")
        num_doc_key     = fm.get("num_doc",  "numDocumentoIdentificacion")

        for r in registros:
            # Filtro por grupo de vida
            if r.get("__grupo") != grupo_id:
                continue

            # Filtro por CUPS (si aplica)
            if cups_set:
                val_cups = str(r.get(cups_field,"") or "").strip().upper()
                # Para otrosServicios el campo puede llamarse diferente
                if archivo == "otrosServicios":
                    val_cups = str(r.get("codServicio","") or r.get("codOtroServicio","") or "").strip().upper()
                cups_norm = {c.upper() for c in cups_set}
                if val_cups not in cups_norm:
                    continue

            # Filtro por nombre (medicamentos con búsqueda de texto)
            if busqueda_nombre:
                val_nombre = _norm(r.get(nombre_field,"") or r.get("nomTecnologiaSalud",""))
                match_nombre = any(term in val_nombre for term in busqueda_nombre)
                if not match_nombre:
                    continue

            # Filtro por finalidad (si se especifica)
            if finalidades:
                val_fin = str(r.get(finalidad_field,"") or "").strip()
                if val_fin not in finalidades:
                    continue

            # Acumular
            if deduplicar:
                pk = (str(r.get(tipo_doc_key,"")).strip(),
                      str(r.get(num_doc_key,"")).strip())
                contados.add(pk)
            else:
                contados.append(1)

        return len(contados)

    # ── Evaluación completa ───────────────────────────────────────────────────
    def evaluar(self, metas: dict) -> dict:
        """
        metas: {programa_id: {actividad_id: int, ...}, ...}
        Retorna resultados completos por programa y actividad
        """
        # Calcular grupos de vida con las fechas reales de atención (igual que Excel)
        self._calcular_grupos()

        cfg_actividades = self.cfg["actividades_base"]
        resultados = {}

        for prog in self.cfg["programas"]:
            prog_id   = prog["id"]
            prog_metas = metas.get(prog_id, {})
            actividades_res = []
            total_exigido   = 0.0
            total_reconocido= 0.0
            total_descuento = 0.0

            # UPC del programa (se calcula como suma_metas_valor / suma_metas_cantidad si se conoce)
            # Si no se conoce, se toma de la nota técnica
            upc_prog = prog_metas.get("__upc", 0)

            for act_id in prog["actividades"]:
                act_cfg = cfg_actividades.get(act_id)
                if not act_cfg:
                    continue

                meta_qty   = int(prog_metas.get(act_id, {}).get("meta", 0) if isinstance(prog_metas.get(act_id), dict) else prog_metas.get(act_id, 0))
                upc_act    = float(prog_metas.get(act_id, {}).get("upc", upc_prog) if isinstance(prog_metas.get(act_id), dict) else upc_prog)
                valor_meta = meta_qty * upc_act

                # Contar ejecutado en RIPS
                ejecutado = self._contar_actividad(act_cfg, prog_id)
                conciliado= min(ejecutado, max(ejecutado, meta_qty))  # conciliado = lo que reportan

                # Discordancia = max(0, meta - conciliado)
                discordancia  = max(0, meta_qty - conciliado)
                valor_reconoc = min(meta_qty, conciliado) * upc_act
                valor_descuento = discordancia * upc_act

                total_exigido    += valor_meta
                total_reconocido += valor_reconoc
                total_descuento  += valor_descuento

                actividades_res.append({
                    "id":          act_id,
                    "descripcion": act_cfg["descripcion"],
                    "meta":        meta_qty,
                    "valor_meta":  valor_meta,
                    "ejecutado":   ejecutado,
                    "conciliado":  conciliado,
                    "discordancia":discordancia,
                    "upc":         upc_act,
                    "valor_descuento": valor_descuento,
                    "valor_reconocido": valor_reconoc,
                    "pct_cumplimiento": (conciliado / meta_qty) if meta_qty > 0 else 1.0
                })

            pct_prog = (total_reconocido / total_exigido) if total_exigido > 0 else 1.0
            resultados[prog_id] = {
                "nombre": prog["nombre"],
                "nombre_acta": prog["nombre_acta"],
                "actividades": actividades_res,
                "total_exigido":    total_exigido,
                "total_reconocido": total_reconocido,
                "total_descuento":  total_descuento,
                "pct_cumplimiento": pct_prog
            }

        return resultados

    # ── Resumen ejecutivo para el acta ────────────────────────────────────────
    def resumen_acta(self, resultados: dict) -> list:
        """Devuelve lista plana de programas con totales para el acta"""
        resumen = []
        for prog_id, r in resultados.items():
            resumen.append({
                "id":           prog_id,
                "nombre":       r["nombre_acta"],
                "exigido":      r["total_exigido"],
                "reconocido":   r["total_reconocido"],
                "descuento":    r["total_descuento"],
                "pct":          r["pct_cumplimiento"]
            })
        return resumen

    # ── Parsear nota técnica (herramienta xlsx) ───────────────────────────────
    @staticmethod
    def parsear_nota_tecnica(path_xlsx: str) -> dict:
        """
        Lee la herramienta xlsx y extrae metas por programa.
        Retorna {prog_id: {act_id: {"meta": N, "upc": X}, "__upc": X}}
        """
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("Instale openpyxl: pip install openpyxl")

        cfg = _load_config()
        wb  = openpyxl.load_workbook(path_xlsx, data_only=True)

        # Buscar hoja A3 COMPLETO
        hoja = None
        for nm in wb.sheetnames:
            if "A3" in nm.upper() or "COMPLETO" in nm.upper():
                hoja = wb[nm]; break
        if not hoja:
            raise ValueError("No se encontró la hoja 'A3 COMPLETO' en el archivo")

        # Mapa de programa → rango de filas (sección)
        FASE_PYMS    = 1
        FASE_MATERNO = 2
        FASE_DI      = 3
        FASE_RCV     = 4

        prog_act_map = {}   # prog_id → {act_id: {meta, valor_meta, upc}}
        fase = FASE_PYMS
        current_prog = None
        current_tipo = None  # tipo de actividad dentro del programa

        # Mapeo de texto descriptivo → actividad_id
        desc_to_act = {_norm(v["descripcion"]): k for k, v in cfg["actividades_base"].items()}

        for row in hoja.iter_rows(min_row=20, values_only=True):
            b = str(row[1] or "").strip()
            c = str(row[2] or "").strip()
            bu = b.upper()

            # Cambio de fase
            if "RUTA INTEGRAL DEMANDA INDUCIDA" in bu and fase < FASE_DI:
                fase = FASE_DI; current_prog = None; continue
            if ("RUTA CEREBROVASCULAR" in bu or "CARDIOVASCULAR" in bu) and fase < FASE_RCV:
                fase = FASE_RCV; current_prog = None; continue
            if ("RUTA MATERNO PERINATAL" in bu or "ATENCION PRECONCEPCIONAL" in bu) and fase < FASE_MATERNO:
                fase = FASE_MATERNO; current_prog = None; continue

            if fase == FASE_PYMS:
                # Detectar sub-sección de programa
                if b and c == "" and "TOTAL" not in bu and "ACTIVIDAD" not in bu:
                    if "PRIMERA INFANCIA" in bu: current_prog = "PRIMERA_INFANCIA"
                    elif "INFANCIA" in bu and "PRIMERA" not in bu: current_prog = "INFANCIA"
                    elif "ADOLESCENCIA" in bu: current_prog = "ADOLESCENCIA"
                    elif "JOVEN" in bu: current_prog = "JOVENES"
                    elif "ADULTEZ" in bu: current_prog = "ADULTEZ"
                    elif "VEJEZ" in bu: current_prog = "VEJEZ"
                elif b and "TOTAL" not in bu and "ACTIVIDAD" not in bu and "DESCRIPCION" not in bu:
                    current_tipo = b
                elif c and c not in ("DESCRIPCION DEL CUPS","") and "TOTAL" not in c and current_prog:
                    # Fila de actividad: c=descripción, row[3]=meta, row[4]=valor_meta, row[8]=upc
                    desc_norm = _norm(c)
                    act_id    = desc_to_act.get(desc_norm)
                    meta_qty  = row[3]
                    val_meta  = row[4]
                    upc_val   = row[8]
                    if act_id and meta_qty is not None:
                        if current_prog not in prog_act_map:
                            prog_act_map[current_prog] = {}
                        try:
                            prog_act_map[current_prog][act_id] = {
                                "meta": int(float(meta_qty)),
                                "upc":  float(upc_val or 0),
                                "valor_meta": float(val_meta or 0)
                            }
                            prog_act_map[current_prog]["__upc"] = float(upc_val or 0)
                        except (ValueError, TypeError):
                            pass

        return prog_act_map
