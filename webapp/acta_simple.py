"""
Generador de acta Excel simplificado para entornos sin el módulo evaluador_3280.py
(usado en Vercel / producción cuando el módulo de escritorio no está disponible)
"""
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _cop(v):
    """Formatea valor como COP"""
    try:
        return f"$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "$ 0,00"

def _pct(v):
    try:
        return f"{float(v)*100:.2f}%"
    except:
        return "0.00%"

# Colores DUSAKAWI
AZUL       = "1F4E79"
AZUL2      = "2E75B6"
VERDE      = "375623"
VERDE2     = "70AD47"
AMARILLO   = "F4B942"
ROJO       = "C00000"
GRIS       = "F2F5F8"
BLANCO     = "FFFFFF"

def _fill(color): return PatternFill("solid", fgColor=color)
def _font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size, name="Calibri")
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _right():  return Alignment(horizontal="right", vertical="center")
def _left():   return Alignment(horizontal="left", vertical="center", wrap_text=True)


def generar_acta_excel(datos_acta: dict, info: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACTA DE EVALUACIÓN"

    # Anchos de columna
    anchos = [4, 34, 18, 18, 18, 12]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 15

    r = 1

    # ── ENCABEZADO ─────────────────────────────────────────
    def hdr_row(label, value, row):
        ws.row_dimensions[row].height = 18
        c = ws.cell(row, 2, label)
        c.font = _font(bold=True, color=BLANCO); c.fill = _fill(AZUL); c.alignment = _left()
        c = ws.cell(row, 3, value)
        c.font = _font(); c.alignment = _left()
        ws.merge_cells(f"C{row}:F{row}")

    # Título
    ws.merge_cells(f"B{r}:F{r}")
    c = ws.cell(r, 2, "DUSAKAWI EPSI")
    c.font = _font(bold=True, color=BLANCO, size=13)
    c.fill = _fill(AZUL); c.alignment = _center()
    ws.row_dimensions[r].height = 28; r += 1

    ws.merge_cells(f"B{r}:F{r}")
    c = ws.cell(r, 2, "ACTA DE EVALUACIÓN – RESOLUCIÓN 3280 DE 2018")
    c.font = _font(bold=True, color=BLANCO, size=11)
    c.fill = _fill(AZUL2); c.alignment = _center()
    ws.row_dimensions[r].height = 22; r += 1

    r += 1

    # Info del acta
    campos = [
        ("Nº Acta", info.get("acta_num","")),
        ("Fecha de evaluación", info.get("fecha_eval","")),
        ("Periodo evaluado", info.get("periodo","")),
        ("Vigencia contrato", info.get("vigencia","")),
        ("Empresa / IPS", info.get("empresa","")),
        ("NIT", info.get("nit","")),
        ("Régimen", info.get("regimen","")),
        ("Municipio", info.get("municipio","")),
        ("Nº Contrato", info.get("num_contrato","")),
        ("Rep. Legal IPS", info.get("rep_ips","")),
        ("Rep. Legal EPSI", info.get("rep_epsi","")),
        ("Coordinador(a)", info.get("coordinador","")),
        ("Evaluador(a)", info.get("evaluador","")),
    ]
    for label, val in campos:
        hdr_row(label, val, r); r += 1

    r += 1

    # ── TABLA DE PROGRAMAS ──────────────────────────────────
    ws.row_dimensions[r].height = 22
    headers = ["#", "PROGRAMA", "VR EXIGIDO", "VR RECONOCIDO", "DESCUENTO", "% CUMPL."]
    for col, h in enumerate(headers, 1):
        c = ws.cell(r, col + 1, h)
        c.font = _font(bold=True, color=BLANCO, size=9)
        c.fill = _fill(AZUL); c.alignment = _center(); c.border = _border()
    r += 1

    programas = datos_acta.get("programas", [])
    for i, prog in enumerate(programas, 1):
        pct_val = prog.get("pct", 0)
        color = VERDE2 if pct_val >= 1.0 else (AMARILLO if pct_val >= 0.9 else ROJO)
        ws.row_dimensions[r].height = 18
        vals = [i, prog.get("nombre",""), _cop(prog.get("exigido",0)),
                _cop(prog.get("reconocido",0)), _cop(prog.get("descuento",0)), _pct(pct_val)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(r, col + 1, val)
            c.font = _font()
            c.alignment = _center() if col in [1, 6] else (_left() if col == 2 else _right())
            c.border = _border()
            if col == 6:
                c.font = _font(bold=True, color=color)
        r += 1

    # Fila de totales
    ws.row_dimensions[r].height = 20
    totales = ["", "TOTAL EJECUCIÓN",
               _cop(datos_acta.get("total_exigido",0)),
               _cop(datos_acta.get("total_reconocido",0)),
               _cop(datos_acta.get("total_descuento",0)), ""]
    for col, val in enumerate(totales, 1):
        c = ws.cell(r, col + 1, val)
        c.font = _font(bold=True, color=BLANCO)
        c.fill = _fill(AZUL); c.alignment = _center() if col != 2 else _left()
        c.border = _border()
    r += 2

    # Descuento final
    ws.row_dimensions[r].height = 18
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2, "TOTAL DESCUENTO (No incluye PAI / Demanda Inducida)")
    c.font = _font(bold=True, color=BLANCO); c.fill = _fill(ROJO); c.alignment = _left()
    c = ws.cell(r, 5, _cop(datos_acta.get("total_descuento",0)))
    c.font = _font(bold=True, color=BLANCO); c.fill = _fill(ROJO); c.alignment = _right()
    r += 2

    # Observaciones
    obs = info.get("observaciones","")
    if obs:
        c = ws.cell(r, 2, "OBSERVACIONES:")
        c.font = _font(bold=True); r += 1
        ws.merge_cells(f"B{r}:F{r}")
        c = ws.cell(r, 2, obs)
        c.font = _font(); c.alignment = _left()
        ws.row_dimensions[r].height = 36; r += 2

    # Firmas
    ws.row_dimensions[r].height = 14
    firmas = [
        ("Rep. Legal IPS", info.get("rep_ips","")),
        ("Coordinador(a)", info.get("coordinador","")),
        ("Evaluador(a)", info.get("evaluador","")),
    ]
    # Líneas de firma
    r_firma = r + 3
    for j, (cargo, nombre) in enumerate(firmas):
        col = 2 + j * 2
        ws.merge_cells(start_row=r_firma, start_column=col, end_row=r_firma, end_column=col+1)
        c = ws.cell(r_firma, col, "_" * 28)
        c.alignment = _center(); c.font = _font()
        ws.merge_cells(start_row=r_firma+1, start_column=col, end_row=r_firma+1, end_column=col+1)
        c = ws.cell(r_firma+1, col, nombre or cargo)
        c.alignment = _center(); c.font = _font(bold=True)
        ws.merge_cells(start_row=r_firma+2, start_column=col, end_row=r_firma+2, end_column=col+1)
        c = ws.cell(r_firma+2, col, cargo)
        c.alignment = _center(); c.font = _font(color="888888")

    # Footer
    r_foot = r_firma + 4
    ws.merge_cells(f"B{r_foot}:F{r_foot}")
    c = ws.cell(r_foot, 2, f"Generado el {datetime.date.today().strftime('%d/%m/%Y')} · Evaluador Res. 3280 v0.1 · DUSAKAWI EPSI")
    c.font = _font(color="888888"); c.alignment = _center()

    wb.save(output_path)
