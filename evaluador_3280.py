#!/usr/bin/env python3
"""
Evaluador Resolución 3280 - DUSAKAWI EPSI
Versión 1.0
Genera actas de evaluación de servicios leyendo la Herramienta de Seguimiento xlsx
"""
import json, os, sys, re, datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ── Rutas por defecto ──────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
STD_JSON   = os.path.join(BASE_DIR, "res3280_standard.json")

# ── Colores / estilos ─────────────────────────────────────────────────────────
AZUL_OSC   = "1F4E79"
AZUL_MED   = "2E75B6"
AZUL_CLAR  = "BDD7EE"
GRIS_ENC   = "D6DCE4"
GRIS_SEC   = "D9D9D9"
VERDE      = "E2EFDA"
AMARILLO   = "FFF2CC"
ROJO       = "FCE4D6"
BLANCO     = "FFFFFF"

def clr(hex_): return PatternFill("solid", fgColor=hex_)
def fnt(bold=False, sz=10, color="000000", name="Calibri"):
    return Font(bold=bold, size=sz, color=color, name=name)
def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(t=True, b=True, l=True, r=True, style="thin"):
    s = Side(style=style) if style else Side(border_style=None)
    return Border(top=s if t else Side(border_style=None),
                  bottom=s if b else Side(border_style=None),
                  left=s if l else Side(border_style=None),
                  right=s if r else Side(border_style=None))

def fmt_pesos(n):
    if n is None: return "$0,00"
    try:
        n = float(n)
        return f"${n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except: return str(n)

def fmt_pct(n):
    if n is None: return "0,00%"
    try: return f"{float(n)*100:.2f}%".replace(".", ",")
    except: return str(n)

# ─────────────────────────────────────────────────────────────────────────────
# PARSER: lee la Herramienta xlsx y extrae totales por programa
# ─────────────────────────────────────────────────────────────────────────────
def parse_herramienta(path):
    """
    Devuelve dict con:
      {
        "identificacion": {...},
        "programas": [
          {
            "id": str, "nombre": str, "nombre_herramienta": str,
            "exigido": float, "reconocido": float, "descuento": float,
            "pct": float, "meta_actividades": int
          }, ...
        ],
        "total_exigido": float, "total_reconocido": float, "total_descuento": float
      }
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # Cargar estándar
    with open(STD_JSON, encoding="utf-8") as f:
        std = json.load(f)
    prog_map = {p["nombre_herramienta"].strip().lower(): p for p in std["programas"]}

    # ── Hoja A3 COMPLETO ──────────────────────────────────────────────────────
    hoja_a3 = None
    for nm in wb.sheetnames:
        if "A3" in nm.upper() or "COMPLETO" in nm.upper():
            hoja_a3 = wb[nm]; break
    if not hoja_a3:
        raise ValueError("No se encontró la hoja 'A3 COMPLETO' en la herramienta")

    identificacion = {}
    LABELS_ID = {
        "CONTRATANTE": "contratante", "NIT": "nit_contratante",
        "CONTRATISTA": "contratista", "TIPO DE CONTRATO": "tipo_contrato",
        "N° DE CONTRATO": "num_contrato", "OBJETO": "objeto",
        "LUGAR DE EJECUCION": "lugar", "VALOR": "valor",
        "VIGENCIA": "vigencia", "POBLACION ASIGNADA:": "poblacion"
    }
    for r in hoja_a3.iter_rows(min_row=1, max_row=25, values_only=True):
        b = str(r[1] or "").strip()
        c = r[2]
        key = LABELS_ID.get(b)
        if key and c is not None:
            if isinstance(c, datetime.datetime):
                c = c.strftime("%d/%m/%Y")
            identificacion[key] = c

    # Leer filas buscando secciones y sus TOTAL
    # Estrategia: fases secuenciales. La herramienta tiene este orden:
    #   1. Rutas PYMS (6 grupos de edad) → filas ~23-186
    #   2. Ruta Materno Perinatal       → filas ~187-263
    #   3. Demanda Inducida             → filas ~264-345
    #   4. Ruta Cerebrovascular/RCV     → filas ~346-462
    # Dentro de cada fase hay sub-secciones; NO deben confundirse entre fases.

    FASE_PYMS   = 1
    FASE_MATERNO= 2
    FASE_DI     = 3
    FASE_RCV    = 4

    resultados = []
    fase = FASE_PYMS
    current_section = None

    di_exigido = di_reconocido = di_descuento = 0.0
    rcv_exigido = rcv_reconocido = rcv_descuento = 0.0
    mat_exigido = mat_reconocido = mat_descuento = 0.0
    pos_exigido = pos_reconocido = pos_descuento = 0.0
    in_pos = False

    for i, row in enumerate(hoja_a3.iter_rows(min_row=20, values_only=True), start=20):
        b = str(row[1] or "").strip()
        c = str(row[2] or "").strip()
        bu = b.upper()

        # ── Detectar cambio de FASE (principal, no sub-sección) ───────────────
        if b and "RUTA INTEGRAL DEMANDA INDUCIDA" in bu and fase < FASE_DI:
            fase = FASE_DI; current_section = "DEMANDA_INDUCIDA"
        elif b and ("RUTA CEREBROVASCULAR" in bu or ("RUTA" in bu and "CARDIOVASCULAR" in bu)) and fase < FASE_RCV:
            fase = FASE_RCV; current_section = "RCV"
        elif b and "ATENCION DEL PARTO" in bu and fase == FASE_MATERNO:
            in_pos = True; current_section = "POSPARTO"
        elif b and ("RUTA MATERNO PERINATAL" in bu or "ATENCION PRECONCEPCIONAL" in bu
                    or "ATENCION CUIDADO PRENATAL" in bu or "ATENCIÓN CUIDADO PRENATAL" in bu
                    or "ATENCIÓN PRECONCEPCIONAL" in bu) and fase < FASE_MATERNO:
            fase = FASE_MATERNO; current_section = "MATERNO_PERINATAL"; in_pos = False

        # ── Detectar sub-sección PYMS (solo en fase PYMS) ────────────────────
        if fase == FASE_PYMS and b and c == "" and "TOTAL" not in bu and "ACTIVIDAD" not in bu:
            if "PRIMERA INFANCIA" in bu:
                current_section = "PRIMERA_INFANCIA"
            elif "INFANCIA" in bu and "PRIMERA" not in bu:
                current_section = "INFANCIA"
            elif "ADOLESCENCIA" in bu:
                current_section = "ADOLESCENCIA"
            elif "JOVEN" in bu:
                current_section = "JOVENES"
            elif "ADULTEZ" in bu:
                current_section = "ADULTEZ"
            elif "VEJEZ" in bu:
                current_section = "VEJEZ"

        # ── Detectar fila TOTAL ───────────────────────────────────────────────
        if c == "TOTAL":
            exigido   = float(row[4] or 0)
            reconocer = float(row[10] or 0)
            descontar = float(row[9] or 0)
            meta      = int(row[3] or 0) if row[3] else 0
            pct_raw   = row[11]
            try: pct = float(pct_raw)
            except: pct = (reconocer / exigido) if exigido else 1.0

            if fase == FASE_PYMS and current_section in (
                    "PRIMERA_INFANCIA","INFANCIA","ADOLESCENCIA","JOVENES","ADULTEZ","VEJEZ"):
                resultados.append({
                    "id": current_section,
                    "exigido": exigido, "reconocido": reconocer,
                    "descuento": descontar, "pct": pct, "meta_actividades": meta
                })
                current_section = None

            elif fase == FASE_MATERNO:
                if in_pos and exigido > 0:
                    pos_exigido += exigido; pos_reconocido += reconocer; pos_descuento += descontar
                    in_pos = False
                elif not in_pos and exigido > 1_000_000:
                    # Gran total de Materno Perinatal
                    mat_exigido = exigido; mat_reconocido = reconocer; mat_descuento = descontar

            elif fase == FASE_DI:
                di_exigido    += exigido
                di_reconocido += reconocer
                di_descuento  += descontar

            elif fase == FASE_RCV:
                # Tomar el acumulado más grande (DM+HTA es el último y mayor)
                if exigido > rcv_exigido:
                    rcv_exigido = exigido; rcv_reconocido = reconocer; rcv_descuento = descontar

    # ── PAI desde hoja LMA ─────────────────────────────────────────────────────
    pai_exigido = 0.0
    hoja_lma = wb["LMA"] if "LMA" in wb.sheetnames else None
    if hoja_lma:
        for row in hoja_lma.iter_rows(min_row=1, max_row=30, values_only=True):
            b = str(row[1] or "").strip()
            if "VACUNACION" in b.upper() or "PAI" in b.upper():
                v = row[1] if isinstance(row[1], (int, float)) else None
                # la fila siguiente tiene el valor
                continue
            if b == "" and row[1] is not None and isinstance(row[1], (int, float)) and pai_exigido == 0:
                # primer valor numérico tras VACUNACION
                pass
        # Buscar por número en la columna B cercano a VACUNACION
        rows_lma = list(hoja_lma.iter_rows(min_row=1, max_row=30, values_only=True))
        for idx, row in enumerate(rows_lma):
            b = str(row[1] or "").strip()
            if "VACUNACION" in b.upper():
                # valor en fila siguiente, col B
                nxt = rows_lma[idx+1] if idx+1 < len(rows_lma) else None
                if nxt and isinstance(nxt[1], (int, float)):
                    pai_exigido = float(nxt[1]); break

    # Armar orden ACTA
    order_ids = ["PRIMERA_INFANCIA","INFANCIA","ADOLESCENCIA","JOVENES","ADULTEZ","VEJEZ",
                 "PAI","MATERNO_PERINATAL","POSPARTO","DEMANDA_INDUCIDA","RCV"]
    nombres_acta = {p["id"]: p["nombre_acta"] for p in std["programas"]}

    prog_result = {r["id"]: r for r in resultados}

    prog_result["PAI"] = {
        "id": "PAI", "exigido": pai_exigido, "reconocido": pai_exigido,
        "descuento": 0.0, "pct": 1.0, "meta_actividades": 0
    }
    prog_result["MATERNO_PERINATAL"] = {
        "id": "MATERNO_PERINATAL", "exigido": mat_exigido, "reconocido": mat_reconocido,
        "descuento": mat_descuento, "pct": (mat_reconocido/mat_exigido) if mat_exigido else 1.0,
        "meta_actividades": 0
    }
    prog_result["POSPARTO"] = {
        "id": "POSPARTO", "exigido": pos_exigido, "reconocido": pos_reconocido,
        "descuento": pos_descuento, "pct": (pos_reconocido/pos_exigido) if pos_exigido else 1.0,
        "meta_actividades": 0
    }
    # Demanda Inducida: descuento = 0 por regla contractual (se reconoce el 100% del exigido)
    prog_result["DEMANDA_INDUCIDA"] = {
        "id": "DEMANDA_INDUCIDA", "exigido": di_exigido, "reconocido": di_exigido,
        "descuento": 0.0, "pct": 1.0,
        "meta_actividades": 0
    }
    prog_result["RCV"] = {
        "id": "RCV", "exigido": rcv_exigido, "reconocido": rcv_reconocido,
        "descuento": rcv_descuento, "pct": (rcv_reconocido/rcv_exigido) if rcv_exigido else 1.0,
        "meta_actividades": 0
    }

    programas_final = []
    for oid in order_ids:
        d = prog_result.get(oid, {
            "id": oid, "exigido": 0.0, "reconocido": 0.0,
            "descuento": 0.0, "pct": 1.0, "meta_actividades": 0
        })
        d["nombre"] = nombres_acta.get(oid, oid)
        programas_final.append(d)

    total_e = sum(p["exigido"]   for p in programas_final)
    total_r = sum(p["reconocido"] for p in programas_final)
    total_d = sum(p["descuento"]  for p in programas_final)

    return {
        "identificacion": identificacion,
        "programas": programas_final,
        "total_exigido": total_e,
        "total_reconocido": total_r,
        "total_descuento": total_d,
    }

# ─────────────────────────────────────────────────────────────────────────────
# GENERADOR DEL ACTA EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def _paint(ws, row, col, value="", bold=False, sz=10, fg="000000", bg=None,
           h="center", v="center", wrap=False, b=True, col_name=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = fnt(bold=bold, sz=sz, color=fg)
    c.alignment = aln(h=h, v=v, wrap=wrap)
    c.border = bdr() if b else Border()
    if bg: c.fill = clr(bg)
    if number_format: c.number_format = number_format
    return c

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def generar_acta_excel(datos, info, output_path):
    """
    datos: resultado de parse_herramienta o dict manual
    info: dict con campos de la GUI {acta_num, fecha_eval, periodo, empresa, municipio,
          nit, regimen, lugar, num_contrato, vigencia, coordinador, evaluador,
          rep_ips, rep_epsi, observaciones, texto_descuento}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACTA"

    # ── Configuración de página ────────────────────────────────────────────────
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "portrait"
    ws.print_area = "A1:J60"
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.7
    ws.page_margins.bottom = 0.7

    # ── Anchos de columna ─────────────────────────────────────────────────────
    COL_W = [2, 6, 6, 8, 8, 8, 8, 7, 7, 8, 0.5]
    # A(2) B(6) C(6) D(8) E(8) F(8) G(8) H(7) I(7) J(8) K(hidden)
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions["K"].width = 0.5
    ws.column_dimensions["K"].hidden = True

    LC = 10  # última columna visible (J)
    row = [1]

    def R():
        return row[0]

    def next_row(h=15):
        ws.row_dimensions[row[0]].height = h
        row[0] += 1

    def paint_row_full(text, bold=False, sz=10, fg="000000", bg=None,
                       h="center", v="center", wrap=False, row_h=15):
        r = R()
        for c in range(1, LC+1):
            _paint(ws, r, c, bg=bg, b=True)
        ws.cell(r, 2).value = text
        ws.cell(r, 2).font  = fnt(bold=bold, sz=sz, color=fg)
        ws.cell(r, 2).alignment = aln(h=h, v=v, wrap=wrap)
        _merge(ws, r, 2, r, LC)
        next_row(row_h)

    # ═══════════════════════════════════════════════════════════════════════════
    # ENCABEZADO - Fila 1
    # ═══════════════════════════════════════════════════════════════════════════
    r = R()
    ws.row_dimensions[r].height = 18
    for c in range(1, LC+1):
        _paint(ws, r, c, bg=AZUL_OSC, b=True)

    # Col A: vacío (logo space)
    ws.cell(r, 1).fill = clr(AZUL_OSC)
    # Título central B-H
    ws.cell(r, 2).value = "PROCESO: GESTIÓN DEL RIESGO EN SALUD\n\nACTA DE EVALUACIÓN DE SERVICIOS"
    ws.cell(r, 2).font  = fnt(bold=True, sz=9, color=BLANCO)
    ws.cell(r, 2).alignment = aln(h="center", v="center", wrap=True)
    _merge(ws, r, 2, r, 8)
    # Código I-J
    codigo_txt = f"CÓDIGO: DR-BC-AP-F-11\nVERSIÓN: 02\nEMISIÓN: 15/06/2022\nVIGENCIA: 27/12/2028\nPÁGINA 1 DE 2"
    ws.cell(r, 9).value = codigo_txt
    ws.cell(r, 9).font  = fnt(sz=7, color=BLANCO)
    ws.cell(r, 9).alignment = aln(h="center", v="center", wrap=True)
    _merge(ws, r, 9, r, LC)
    ws.row_dimensions[r].height = 55
    next_row(55)

    # ── Línea 2: separador ────────────────────────────────────────────────────
    r = R()
    ws.row_dimensions[r].height = 3
    for c in range(1, LC+1):
        _paint(ws, r, c, bg=AZUL_MED, b=False)
    next_row(3)

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLA DE IDENTIFICACIÓN
    # ═══════════════════════════════════════════════════════════════════════════
    def id_row(lbl1, val1, lbl2="", val2="", merge_val1=3, col_lbl2=7, row_h=14):
        r = R()
        ws.row_dimensions[r].height = row_h
        # Etiqueta 1 (B-C)
        for c in range(1, LC+1): _paint(ws, r, c, bg=GRIS_ENC, b=True)
        ws.cell(r, 2).value = lbl1; ws.cell(r, 2).font = fnt(bold=True, sz=9)
        ws.cell(r, 2).alignment = aln(h="left")
        _merge(ws, r, 2, r, 2+merge_val1-1)
        # Valor 1
        v1_col = 2+merge_val1
        ws.cell(r, v1_col).value = val1; ws.cell(r, v1_col).font = fnt(sz=9)
        ws.cell(r, v1_col).alignment = aln(h="left")
        _merge(ws, r, v1_col, r, col_lbl2-1)
        # Etiqueta 2
        if lbl2:
            ws.cell(r, col_lbl2).value = lbl2; ws.cell(r, col_lbl2).font = fnt(bold=True, sz=9)
            ws.cell(r, col_lbl2).alignment = aln(h="left")
            _merge(ws, r, col_lbl2, r, col_lbl2)
            # Valor 2
            ws.cell(r, col_lbl2+1).value = val2; ws.cell(r, col_lbl2+1).font = fnt(sz=9)
            ws.cell(r, col_lbl2+1).alignment = aln(h="left")
            _merge(ws, r, col_lbl2+1, r, LC)
        next_row(row_h)

    periodo = info.get("periodo","")
    empresa = info.get("empresa","")
    nit     = info.get("nit","")
    regimen = info.get("regimen","")
    municipio = info.get("municipio","")
    lugar   = info.get("lugar","")
    num_contrato = info.get("num_contrato","")
    vigencia = info.get("vigencia","")
    fecha_eval = info.get("fecha_eval","")
    acta_num = info.get("acta_num","")

    id_row("ACTA Nº", acta_num, "FECHA DE EVALUACIÓN", fecha_eval)
    id_row("VIGENCIA DEL CONTRATO:", vigencia, "PERIODO EVALUADO", periodo)
    id_row("NIT", nit, "EMPRESA", empresa)
    id_row("LUGAR", lugar, "MUNICIPIO", municipio)
    id_row("REGIMEN", regimen, "Nº CONTRATO", num_contrato)

    # ── PUNTOS A TRATAR ───────────────────────────────────────────────────────
    r = R(); ws.row_dimensions[r].height = 14
    for c in range(1, LC+1): _paint(ws, r, c, bg=GRIS_SEC, b=True)
    ws.cell(r, 2).value = "PUNTOS A TRATAR:"
    ws.cell(r, 2).font  = fnt(bold=True, sz=9)
    ws.cell(r, 2).alignment = aln(h="left")
    _merge(ws, r, 2, r, LC)
    next_row(14)

    PUNTOS = ("Revisión y confrontación de resultados del seguimiento a las actividades de "
              "promoción y mantenimiento de la salud y atención primaria, reportadas según "
              "Anexos contractuales N°. 11 y 12  por la coordinación de baja complejidad.")
    paint_row_full(PUNTOS, h="left", wrap=True, row_h=28)

    r = R(); ws.row_dimensions[r].height = 14
    for c in range(1, LC+1): _paint(ws, r, c, bg=GRIS_SEC, b=True)
    ws.cell(r, 2).value = "OBJETIVO:"
    ws.cell(r, 2).font  = fnt(bold=True, sz=9)
    ws.cell(r, 2).alignment = aln(h="left")
    _merge(ws, r, 2, r, LC)
    next_row(14)

    OBJETIVO = ("Verificar la ejecución de la efectiva prestación de los servicios, cumplimiento "
                "de metas reflejadas en los RIPS cargados y facturas radicadas por parte de los "
                "prestadores de la red de atención primaria bajo la modalidad de cápita, según "
                "los anexos 10, 11 y 12.")
    paint_row_full(OBJETIVO, h="left", wrap=True, row_h=28)

    r = R(); ws.row_dimensions[r].height = 14
    for c in range(1, LC+1): _paint(ws, r, c, bg=GRIS_SEC, b=True)
    ws.cell(r, 2).value = "DESARROLLO Y CONCLUSIONES"
    ws.cell(r, 2).font  = fnt(bold=True, sz=9)
    ws.cell(r, 2).alignment = aln(h="left")
    _merge(ws, r, 2, r, LC)
    next_row(14)

    DESARROLLO = (f"Se realiza la evaluación de la ejecución de las actividades contenidas en las "
                  f"rutas de atención en salud correspondientes a los meses de {periodo}. "
                  f"En donde se evidencia lo relacionado en la gráfica 1.")
    paint_row_full(DESARROLLO, h="left", wrap=True, row_h=28)

    # ── Espacio gráfica (se inserta el gráfico aquí) ─────────────────────────
    chart_row_start = R()
    for _ in range(14):
        r = R()
        ws.row_dimensions[r].height = 14
        for c in range(1, LC+1): _paint(ws, r, c, bg=None, b=False)
        next_row(14)

    # ── Texto pre-tabla ───────────────────────────────────────────────────────
    PRE_TABLA = ("Una vez verificada la ejecución de las actividades de promoción y mantenimiento "
                 "de la salud, se establece la siguiente relación de descuento por incumplimiento "
                 "de metas, las cuales se desglosan por programa contratado en la siguiente tabla:")
    paint_row_full(PRE_TABLA, h="left", wrap=True, row_h=28)

    # ═══════════════════════════════════════════════════════════════════════════
    # TABLA DE PROGRAMAS
    # ═══════════════════════════════════════════════════════════════════════════
    # Encabezado de tabla
    r = R(); ws.row_dimensions[r].height = 30
    headers = ["Programa", "", "", "", "", "Vr Exigido", "Vr Reconocido", "Descuento", "% Cumpl.", ""]
    bg_h = AZUL_MED
    for ci, h_text in enumerate(headers, start=1):
        _paint(ws, r, ci, value=h_text, bold=True, sz=9, fg=BLANCO, bg=bg_h,
               h="center", v="center", wrap=True)
    ws.cell(r, 2).value = "Programa"
    _merge(ws, r, 2, r, 5)
    next_row(30)

    programas = datos.get("programas", [])
    tabla_data_rows = []

    # Color de fila alterno
    for idx, prog in enumerate(programas):
        r = R(); ws.row_dimensions[r].height = 20
        bg_row = BLANCO if idx % 2 == 0 else AZUL_CLAR
        for c in range(1, LC+1): _paint(ws, r, c, bg=bg_row, b=True)

        exigido   = prog.get("exigido", 0) or 0
        reconocido= prog.get("reconocido", 0) or 0
        descuento = prog.get("descuento", 0) or 0
        pct       = prog.get("pct", 1.0) or 1.0

        ws.cell(r, 2).value = prog.get("nombre", "")
        ws.cell(r, 2).font  = fnt(sz=8, bold=False)
        ws.cell(r, 2).alignment = aln(h="left", v="center", wrap=True)
        _merge(ws, r, 2, r, 5)

        ws.cell(r, 6).value  = exigido;    ws.cell(r, 6).number_format = '#,##0.00'
        ws.cell(r, 7).value  = reconocido; ws.cell(r, 7).number_format = '#,##0.00'
        ws.cell(r, 8).value  = descuento;  ws.cell(r, 8).number_format = '#,##0.00'
        ws.cell(r, 9).value  = pct;        ws.cell(r, 9).number_format = '0.00%'
        # Col K (hidden) para gráfica — % cumplimiento
        ws.cell(r, 11).value = pct
        ws.cell(r, 11).font  = fnt(color=BLANCO, sz=7)

        for c in [6, 7, 8, 9]:
            ws.cell(r, c).font      = fnt(sz=8)
            ws.cell(r, c).alignment = aln(h="center", v="center")

        # Color de celda pct según umbral
        if pct >= 1.0:
            ws.cell(r, 9).fill = clr(VERDE)
        elif pct >= 0.90:
            ws.cell(r, 9).fill = clr(AMARILLO)
        else:
            ws.cell(r, 9).fill = clr(ROJO)

        tabla_data_rows.append(r)
        next_row(20)

    # Totales
    total_e = datos.get("total_exigido", 0) or 0
    total_r = datos.get("total_reconocido", 0) or 0
    total_d = datos.get("total_descuento", 0) or 0

    r = R(); ws.row_dimensions[r].height = 18
    for c in range(1, LC+1): _paint(ws, r, c, bg=AZUL_MED, b=True)
    ws.cell(r, 2).value = "TOTAL EJECUCIÓN"
    ws.cell(r, 2).font  = fnt(bold=True, sz=9, color=BLANCO)
    ws.cell(r, 2).alignment = aln(h="left")
    _merge(ws, r, 2, r, 5)
    for ci, val in [(6, total_e), (7, total_r), (8, total_d)]:
        ws.cell(r, ci).value = val
        ws.cell(r, ci).number_format = '#,##0.00'
        ws.cell(r, ci).font = fnt(bold=True, sz=9, color=BLANCO)
        ws.cell(r, ci).alignment = aln(h="center")
    next_row(18)

    # PAI excluido del descuento total
    pai_d = next((p.get("descuento",0) for p in programas if p.get("id")=="PAI"), 0)
    total_d_sin_pai = total_d - (pai_d or 0)

    r = R(); ws.row_dimensions[r].height = 18
    for c in range(1, LC+1): _paint(ws, r, c, bg=AZUL_OSC, b=True)
    ws.cell(r, 2).value = "TOTAL DESCUENTO (No Incluye PAI)"
    ws.cell(r, 2).font  = fnt(bold=True, sz=9, color=BLANCO)
    ws.cell(r, 2).alignment = aln(h="left")
    _merge(ws, r, 2, r, 7)
    ws.cell(r, 8).value = total_d_sin_pai
    ws.cell(r, 8).number_format = '#,##0.00'
    ws.cell(r, 8).font = fnt(bold=True, sz=10, color=AMARILLO)
    ws.cell(r, 8).alignment = aln(h="center")
    next_row(18)

    # ═══════════════════════════════════════════════════════════════════════════
    # GRÁFICA DE BARRAS — % Cumplimiento por Programa
    # ═══════════════════════════════════════════════════════════════════════════
    if tabla_data_rows:
        chart = BarChart()
        chart.type        = "col"
        chart.grouping    = "clustered"
        chart.title       = "% Cumplimiento por Programa - Res. 3280"
        chart.y_axis.title = "% Cumplimiento"
        chart.x_axis.title = "Programa"
        chart.style       = 10
        chart.width       = 22
        chart.height      = 13

        min_r = tabla_data_rows[0]
        max_r = tabla_data_rows[-1]

        data_ref = Reference(ws, min_col=11, max_col=11, min_row=min_r, max_row=max_r)
        cats_ref = Reference(ws, min_col=2,  max_col=2,  min_row=min_r, max_row=max_r)
        chart.add_data(data_ref)
        chart.set_categories(cats_ref)
        chart.series[0].title = openpyxl.chart.series.SeriesLabel(v="% Cumplimiento")
        chart.series[0].graphicalProperties.solidFill = AZUL_MED

        # Línea de meta 100%
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.chart import LineChart
        line = LineChart()
        line.grouping = "standard"
        # No podemos hacer serie de referencia sin datos, así que solo BarChart
        chart.y_axis.scaling.max = 1.2
        chart.y_axis.scaling.min = 0.0
        chart.y_axis.numFmt = "0%"
        chart.plot_area.dTable = None

        ws.add_chart(chart, f"B{chart_row_start}")

    # ═══════════════════════════════════════════════════════════════════════════
    # PÁGINA 2 — Textos legales y firmas
    # ═══════════════════════════════════════════════════════════════════════════
    # Separador de página (fila en blanco)
    r = R(); ws.row_dimensions[r].height = 6
    for c in range(1, LC+1): _paint(ws, r, c, bg=None, b=False)
    next_row(6)

    # Encabezado página 2
    r = R(); ws.row_dimensions[r].height = 55
    for c in range(1, LC+1): _paint(ws, r, c, bg=AZUL_OSC, b=True)
    ws.cell(r, 2).value = "PROCESO: GESTIÓN DEL RIESGO EN SALUD\n\nACTA DE EVALUACIÓN DE SERVICIOS"
    ws.cell(r, 2).font  = fnt(bold=True, sz=9, color=BLANCO)
    ws.cell(r, 2).alignment = aln(h="center", v="center", wrap=True)
    _merge(ws, r, 2, r, 8)
    ws.cell(r, 9).value = "CÓDIGO: DR-BC-AP-F-11\nVERSIÓN: 02\nEMISIÓN: 15/06/2022\nVIGENCIA: 27/12/2022\nPÁGINA 2 DE 2"
    ws.cell(r, 9).font  = fnt(sz=7, color=BLANCO)
    ws.cell(r, 9).alignment = aln(h="center", v="center", wrap=True)
    _merge(ws, r, 9, r, LC)
    next_row(55)

    # Texto de proyección de descuento
    desc_txt = info.get("texto_descuento", "")
    if not desc_txt:
        desc_txt = (f"En línea con lo anterior la proyección final de descuento es de "
                    f"{fmt_pesos(total_d_sin_pai)} se verifica la ejecución de las actividades "
                    f"de atención primaria, correspondientes a la vigencia antes mencionada, "
                    f"se aclara que se debe cumplir con mínimo el 100% de las actividades "
                    f"programadas para los servicios de promoción y mantenimiento y ruta materno.")
    paint_row_full(desc_txt, h="left", wrap=True, row_h=35)

    # OBSERVACIONES
    r = R(); ws.row_dimensions[r].height = 14
    for c in range(1, LC+1): _paint(ws, r, c, bg=GRIS_SEC, b=True)
    ws.cell(r, 2).value = "OBSERVACIONES:"
    ws.cell(r, 2).font  = fnt(bold=True, sz=9)
    ws.cell(r, 2).alignment = aln(h="left")
    _merge(ws, r, 2, r, LC)
    next_row(14)

    obs = info.get("observaciones", "")
    paint_row_full(obs, h="left", wrap=True, row_h=28)

    # Textos legales
    NOTAS = [
        ('NOTA: Para efectos del cumplimiento de lo establecido por las partes en el contrato '
         'suscrito para la vigencia y, en consecuencia de ello el seguimiento al cumplimiento '
         'de la Programación de las actividades y procedimientos de PROTECCIÓN ESPECIFICA y '
         'DETECCIÓN TEMPRANA, se tendrá en cuenta lo establecido en la normatividad vigente '
         'Decreto 780 de 2016, las Resoluciones 3100 de 2019, Resolución 4003 de 2008, '
         'Resolución 2626 de 2019, Resolución 3202 de 2016, Resolución 3280 de agosto 2018 '
         'y Resolución 3253 de 2009, 202 del 2020, así como lo dispuesto en el ANEXO N° 14 '
         'del contrato. "PROCEDIMIENTO SEGUIMIENTO A CONTRATOS".'),
        ('En tal sentido, se realizará el SEGUIMIENTO Y EVALUACIÒN mensual por parte de '
         'DUSAKAWI EPSI y de manera trimestral entre las Partes conjuntamente las respectivas '
         'REUNIONES DE GESTIÓN COMPARTIDA. En el evento en que, de las actividades de seguimiento '
         'y evaluación, así como de las reuniones de gestión compartida se determine el '
         'incumplimiento de una o más metas, el prestador de servicios deberá elaborar presentar '
         'ante la EPSI dentro de los cinco (5) días hábiles siguientes a la realización de la '
         'reunión de gestión compartida un Plan de Mejoramiento que defina acciones y estrategias '
         'que le permita lograr el mayor porcentaje de cumplimiento de las metas incumplidas y '
         'las sucesivas.'),
        ('Los resultados de la evaluación del cumplimiento de metas registrados en el Acta de '
         'Gestión Compartida que se suscribirá por cada trimestre que generen proyección de '
         'descuento en contra del prestador será registrada contablemente. Si posterior a la '
         'evaluación del Plan de Mejoramiento suscrito por el prestador y en todo caso del '
         'trimestre inmediatamente siguiente al de aquel en que se han generado proyección de '
         'descuentos el prestador de servicios no lograre subir la meta, la suba parcialmente '
         'o la cumpla en su totalidad, respectivamente se procederá a materializar el valor '
         'proyectado del descuento en contra del prestador.'),
        ('Lo anterior no impide que durante lo que reste de la vigencia del contrato el prestador '
         'pueda alcanzar el cumplimiento de las metas de los meses anteriores, siempre que sea '
         'procedente su acumulación, caso en el cual, se procederá a reconocer el valor que '
         'corresponda y se registrará contablemente como una nota crédito a favor del prestador, '
         'proporcional a las actividades realizadas.')
    ]
    for nota in NOTAS:
        paint_row_full(nota, h="left", wrap=True, row_h=38)

    # ── FIRMAS ────────────────────────────────────────────────────────────────
    def firma_row(nombre1, cargo1, nombre2, cargo2, row_h=16):
        # Nombre
        r = R(); ws.row_dimensions[r].height = row_h
        for c in range(1, LC+1): _paint(ws, r, c, bg=None, b=False)
        ws.cell(r, 2).value = nombre1; ws.cell(r, 2).font = fnt(bold=True, sz=9)
        ws.cell(r, 2).alignment = aln(h="center"); _merge(ws, r, 2, r, 4)
        ws.cell(r, 7).value = nombre2; ws.cell(r, 7).font = fnt(bold=True, sz=9)
        ws.cell(r, 7).alignment = aln(h="center"); _merge(ws, r, 7, r, LC)
        next_row(row_h)
        # Cargo
        r = R(); ws.row_dimensions[r].height = row_h
        for c in range(1, LC+1): _paint(ws, r, c, bg=None, b=False)
        ws.cell(r, 2).value = cargo1; ws.cell(r, 2).font = fnt(sz=8)
        ws.cell(r, 2).alignment = aln(h="center"); _merge(ws, r, 2, r, 4)
        ws.cell(r, 7).value = cargo2; ws.cell(r, 7).font = fnt(sz=8)
        ws.cell(r, 7).alignment = aln(h="center"); _merge(ws, r, 7, r, LC)
        # Líneas de firma
        for c in [2, 3, 4]: ws.cell(r, c).border = bdr(t=False,b=True,l=False,r=False)
        for c in [7, 8, 9, 10]: ws.cell(r, c).border = bdr(t=False,b=True,l=False,r=False)
        next_row(row_h)

    # Espacio para firmas
    r = R(); ws.row_dimensions[r].height = 20
    for c in range(1, LC+1): _paint(ws, r, c, bg=None, b=False)
    next_row(20)

    rep_ips  = info.get("rep_ips",  "REPRESENTANTE LEGAL IPS")
    rep_epsi = info.get("rep_epsi", "JAVIER CLAVIJO FRANCO")
    coord    = info.get("coordinador", "COORDINADOR(A) DE BAJA COMPLEJIDAD")
    evaluador= info.get("evaluador",   "FUNCIONARIO QUIEN REALIZA LA EVALUACIÓN")

    firma_row(rep_ips, "REPRESENTANTE LEGAL", rep_epsi, "REPRESENTANTE LEGAL EPSI")

    r = R(); ws.row_dimensions[r].height = 20
    for c in range(1, LC+1): _paint(ws, r, c, bg=None, b=False)
    next_row(20)

    firma_row(coord, "COORDINADOR(A) DE BAJA COMPLEJIDAD", evaluador, "FUNCIONARIO QUIEN REALIZA LA EVALUACIÓN")

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# INTERFAZ GRÁFICA
# ─────────────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Evaluador Res. 3280 – DUSAKAWI EPSI  v1.0")
        self.geometry("1000x680")
        self.resizable(True, True)
        self.configure(bg="#F0F4F8")

        # Carga el estándar
        with open(STD_JSON, encoding="utf-8") as f:
            self.standard = json.load(f)

        self._herramienta_path = tk.StringVar()
        self._parsed_data = None
        self._prog_vars = {}  # id -> {"exigido": DoubleVar, "reconocido": DoubleVar, "descuento": DoubleVar}

        self._build_ui()

    def _build_ui(self):
        STYLE_BG  = "#F0F4F8"
        STYLE_HDR = "#1F4E79"
        # Header
        hdr = tk.Frame(self, bg=STYLE_HDR, height=50)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Evaluador Resolución 3280 – DUSAKAWI EPSI",
                 font=("Calibri", 14, "bold"), fg="white", bg=STYLE_HDR
                 ).pack(side="left", padx=14, pady=10)
        tk.Label(hdr, text="v1.0", font=("Calibri", 9), fg="#BDD7EE", bg=STYLE_HDR
                 ).pack(side="right", padx=14)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=8)

        tab1 = tk.Frame(nb, bg=STYLE_BG)
        tab2 = tk.Frame(nb, bg=STYLE_BG)
        tab3 = tk.Frame(nb, bg=STYLE_BG)
        nb.add(tab1, text="  1. Herramienta & Datos  ")
        nb.add(tab2, text="  2. Programas  ")
        nb.add(tab3, text="  3. Generar Acta  ")

        self._build_tab1(tab1)
        self._build_tab2(tab2)
        self._build_tab3(tab3)

    # ── TAB 1: Selección herramienta + datos generales ────────────────────────
    def _build_tab1(self, parent):
        # Sección archivo
        frm_file = tk.LabelFrame(parent, text="Herramienta de Seguimiento (xlsx)",
                                 font=("Calibri", 10, "bold"), bg="#F0F4F8",
                                 fg="#1F4E79", padx=10, pady=8)
        frm_file.pack(fill="x", padx=14, pady=(12, 6))

        tk.Entry(frm_file, textvariable=self._herramienta_path, width=65,
                 font=("Calibri", 9)).pack(side="left", padx=4)
        tk.Button(frm_file, text="Seleccionar...", command=self._select_herramienta,
                  bg="#2E75B6", fg="white", font=("Calibri", 9, "bold"),
                  relief="flat", padx=10).pack(side="left", padx=4)
        tk.Button(frm_file, text="⟳ Leer Datos", command=self._load_herramienta,
                  bg="#375623", fg="white", font=("Calibri", 9, "bold"),
                  relief="flat", padx=10).pack(side="left", padx=4)

        # Datos generales
        frm_id = tk.LabelFrame(parent, text="Datos de Identificación del Acta",
                                font=("Calibri", 10, "bold"), bg="#F0F4F8",
                                fg="#1F4E79", padx=12, pady=10)
        frm_id.pack(fill="both", expand=True, padx=14, pady=6)

        campos = [
            ("Acta Nº",           "acta_num",      "PMS-44430-2026-XX-X"),
            ("Fecha de Evaluación","fecha_eval",     datetime.date.today().strftime("%d/%m/%Y")),
            ("Periodo Evaluado",   "periodo",        "MAR"),
            ("Vigencia del Contrato","vigencia",     "01/03/2026- 31/12/2026"),
            ("Empresa / IPS",      "empresa",        "IPSI KARAQUITA"),
            ("NIT",                "nit",            "900141404"),
            ("Régimen",            "regimen",        "SUBSIDIADO"),
            ("Municipio",          "municipio",      "MAICAO"),
            ("Lugar",              "lugar",          "VALLEDUPAR"),
            ("Nº Contrato",        "num_contrato",   "PMS-44430-2026-16"),
            ("Rep. Legal IPS",     "rep_ips",        ""),
            ("Rep. Legal EPSI",    "rep_epsi",       "JAVIER CLAVIJO FRANCO"),
            ("Coordinador BC",     "coordinador",    "DIANA CASTILLA BARRAZA"),
            ("Evaluador",          "evaluador",      "JESUS VANEGAS FONTALVO"),
        ]
        self._id_vars = {}
        for i, (lbl, key, default) in enumerate(campos):
            r, c = divmod(i, 2)
            tk.Label(frm_id, text=lbl+":", font=("Calibri", 9), bg="#F0F4F8",
                     anchor="e").grid(row=r, column=c*3, sticky="e", padx=(8,4), pady=3)
            var = tk.StringVar(value=default)
            self._id_vars[key] = var
            tk.Entry(frm_id, textvariable=var, width=28,
                     font=("Calibri", 9)).grid(row=r, column=c*3+1, sticky="w", padx=(0,16), pady=3)

        # Observaciones
        tk.Label(frm_id, text="Observaciones:", font=("Calibri", 9, "bold"),
                 bg="#F0F4F8").grid(row=8, column=0, sticky="nw", padx=(8,4), pady=3)
        self._obs_text = tk.Text(frm_id, height=4, width=80, font=("Calibri", 9),
                                 wrap="word")
        self._obs_text.grid(row=8, column=1, columnspan=5, sticky="ew", padx=(0,8), pady=3)

    # ── TAB 2: Tabla de programas (editable) ──────────────────────────────────
    def _build_tab2(self, parent):
        tk.Label(parent, text="Valores por Programa  (se autocompletan al leer la Herramienta)",
                 font=("Calibri", 10), bg="#F0F4F8", fg="#1F4E79").pack(pady=(10,4), anchor="w", padx=14)

        cols = ("programa", "exigido", "reconocido", "descuento", "pct_cumpl")
        headers = ("Programa", "Vr Exigido", "Vr Reconocido", "Descuento", "% Cumpl.")

        frm = tk.Frame(parent, bg="#F0F4F8")
        frm.pack(fill="both", expand=True, padx=14, pady=4)

        self._tree = ttk.Treeview(frm, columns=cols, show="headings", height=14)
        widths = [260, 120, 120, 120, 80]
        for col, hdr, w in zip(cols, headers, widths):
            self._tree.heading(col, text=hdr)
            self._tree.column(col, width=w, anchor="center" if col != "programa" else "w")
        self._tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(frm, orient="vertical", command=self._tree.yview)
        sb.pack(side="right", fill="y")
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.bind("<Double-1>", self._on_tree_edit)

        # Totales
        frm_tot = tk.Frame(parent, bg="#D6DCE4")
        frm_tot.pack(fill="x", padx=14, pady=4)
        self._lbl_total_e = tk.Label(frm_tot, text="Total Exigido: $0", font=("Calibri", 10, "bold"), bg="#D6DCE4")
        self._lbl_total_e.pack(side="left", padx=16, pady=4)
        self._lbl_total_r = tk.Label(frm_tot, text="Total Reconocido: $0", font=("Calibri", 10, "bold"), bg="#D6DCE4")
        self._lbl_total_r.pack(side="left", padx=16)
        self._lbl_total_d = tk.Label(frm_tot, text="Total Descuento: $0", font=("Calibri", 10, "bold"), bg="#D6DCE4", fg="#C00000")
        self._lbl_total_d.pack(side="left", padx=16)

        tk.Label(parent, text="Doble clic en una fila para editar los valores manualmente.",
                 font=("Calibri", 8), fg="gray", bg="#F0F4F8").pack(pady=2, anchor="w", padx=14)

    # ── TAB 3: Generar acta ────────────────────────────────────────────────────
    def _build_tab3(self, parent):
        frm = tk.Frame(parent, bg="#F0F4F8")
        frm.pack(fill="both", expand=True, padx=20, pady=20)

        tk.Label(frm, text="Texto de Proyección de Descuento (opcional, se auto-genera):",
                 font=("Calibri", 10), bg="#F0F4F8").pack(anchor="w", pady=(0,4))
        self._desc_text = tk.Text(frm, height=5, width=90, font=("Calibri", 9), wrap="word")
        self._desc_text.pack(fill="x", pady=(0,16))

        tk.Label(frm, text="Ruta de salida del archivo Excel:",
                 font=("Calibri", 10), bg="#F0F4F8").pack(anchor="w")
        frm2 = tk.Frame(frm, bg="#F0F4F8")
        frm2.pack(fill="x", pady=4)
        self._out_var = tk.StringVar(value=os.path.join(
            os.path.expanduser("~/Desktop"), "ACTA_EVALUACION.xlsx"))
        tk.Entry(frm2, textvariable=self._out_var, width=60,
                 font=("Calibri", 9)).pack(side="left", padx=4)
        tk.Button(frm2, text="Guardar como...", command=self._pick_output,
                  bg="#2E75B6", fg="white", font=("Calibri", 9, "bold"),
                  relief="flat", padx=10).pack(side="left")

        tk.Button(frm, text="GENERAR ACTA EXCEL", command=self._generar,
                  bg="#1F4E79", fg="white", font=("Calibri", 14, "bold"),
                  relief="flat", padx=30, pady=12).pack(pady=30)

        self._status_var = tk.StringVar(value="")
        tk.Label(frm, textvariable=self._status_var, font=("Calibri", 9),
                 bg="#F0F4F8", fg="#375623", wraplength=700, justify="left").pack()

    # ── Acciones ──────────────────────────────────────────────────────────────
    def _select_herramienta(self):
        path = filedialog.askopenfilename(
            title="Seleccionar Herramienta de Seguimiento",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if path:
            self._herramienta_path.set(path)

    def _load_herramienta(self):
        path = self._herramienta_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", "Seleccione un archivo de Herramienta válido.")
            return
        try:
            data = parse_herramienta(path)
            self._parsed_data = data
            # Rellenar identificación
            ident = data.get("identificacion", {})
            for key, var in self._id_vars.items():
                if key in ident and not var.get():
                    var.set(str(ident[key]))
            self._populate_tree(data)
            messagebox.showinfo("OK", "Herramienta cargada correctamente.\n"
                                      "Revise la Pestaña 2 para verificar los valores.")
        except Exception as e:
            messagebox.showerror("Error al leer herramienta", str(e))

    def _populate_tree(self, data):
        for item in self._tree.get_children():
            self._tree.delete(item)
        for prog in data.get("programas", []):
            e = prog.get("exigido", 0) or 0
            r = prog.get("reconocido", 0) or 0
            d = prog.get("descuento", 0) or 0
            pct = prog.get("pct", 1.0) or 1.0
            tag = "verde" if pct >= 1.0 else ("amarillo" if pct >= 0.90 else "rojo")
            self._tree.insert("", "end", iid=prog["id"],
                              values=(prog["nombre"],
                                      f"{e:,.2f}", f"{r:,.2f}", f"{d:,.2f}",
                                      f"{pct*100:.2f}%"),
                              tags=(tag,))
        self._tree.tag_configure("verde",    background="#E2EFDA")
        self._tree.tag_configure("amarillo", background="#FFF2CC")
        self._tree.tag_configure("rojo",     background="#FCE4D6")

        te = data.get("total_exigido", 0) or 0
        tr = data.get("total_reconocido", 0) or 0
        td = data.get("total_descuento", 0) or 0
        self._lbl_total_e.config(text=f"Total Exigido: {fmt_pesos(te)}")
        self._lbl_total_r.config(text=f"Total Reconocido: {fmt_pesos(tr)}")
        self._lbl_total_d.config(text=f"Total Descuento: {fmt_pesos(td)}")

    def _on_tree_edit(self, event):
        item = self._tree.selection()
        if not item: return
        iid = item[0]
        vals = self._tree.item(iid, "values")
        # Diálogo de edición
        dlg = tk.Toplevel(self)
        dlg.title("Editar valores")
        dlg.geometry("360x200")
        dlg.grab_set()
        labels = ["Programa", "Vr Exigido", "Vr Reconocido", "Descuento"]
        vars_  = []
        for i, (lbl, val) in enumerate(zip(labels, vals[:4])):
            tk.Label(dlg, text=lbl+":").grid(row=i, column=0, sticky="e", padx=8, pady=4)
            v = tk.StringVar(value=val.replace(",","") if i > 0 else val)
            tk.Entry(dlg, textvariable=v, width=22,
                     state="disabled" if i==0 else "normal").grid(row=i, column=1, sticky="w")
            vars_.append(v)
        def save():
            try:
                e = float(vars_[1].get()); r = float(vars_[2].get()); d = float(vars_[3].get())
                pct = r/e if e else 1.0
                tag = "verde" if pct >= 1.0 else ("amarillo" if pct >= 0.90 else "rojo")
                self._tree.item(iid, values=(vals[0], f"{e:,.2f}", f"{r:,.2f}", f"{d:,.2f}", f"{pct*100:.2f}%"), tags=(tag,))
                if self._parsed_data:
                    for prog in self._parsed_data["programas"]:
                        if prog["id"] == iid:
                            prog["exigido"] = e; prog["reconocido"] = r
                            prog["descuento"] = d; prog["pct"] = pct
                # Recalcular totales
                if self._parsed_data:
                    progs = self._parsed_data["programas"]
                    te = sum(p.get("exigido",0) or 0 for p in progs)
                    tr = sum(p.get("reconocido",0) or 0 for p in progs)
                    td = sum(p.get("descuento",0) or 0 for p in progs)
                    self._parsed_data["total_exigido"] = te
                    self._parsed_data["total_reconocido"] = tr
                    self._parsed_data["total_descuento"] = td
                    self._lbl_total_e.config(text=f"Total Exigido: {fmt_pesos(te)}")
                    self._lbl_total_r.config(text=f"Total Reconocido: {fmt_pesos(tr)}")
                    self._lbl_total_d.config(text=f"Total Descuento: {fmt_pesos(td)}")
                dlg.destroy()
            except ValueError:
                messagebox.showerror("Error", "Ingrese valores numéricos válidos.", parent=dlg)
        tk.Button(dlg, text="Guardar", command=save, bg="#2E75B6", fg="white",
                  font=("Calibri", 10, "bold"), relief="flat").grid(row=4, column=0, columnspan=2, pady=12)

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title="Guardar Acta como...",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="ACTA_EVALUACION.xlsx")
        if path: self._out_var.set(path)

    def _build_info_dict(self):
        info = {k: v.get() for k, v in self._id_vars.items()}
        info["observaciones"]   = self._obs_text.get("1.0", "end").strip()
        info["texto_descuento"] = self._desc_text.get("1.0", "end").strip()
        return info

    def _generar(self):
        if not self._parsed_data:
            if not messagebox.askyesno("Sin datos",
               "No se cargó ninguna herramienta.\n¿Continuar con valores en cero?"):
                return
            # Crear estructura vacía
            with open(STD_JSON, encoding="utf-8") as f:
                std = json.load(f)
            progs = [{"id": p["id"], "nombre": p["nombre_acta"],
                      "exigido":0, "reconocido":0, "descuento":0, "pct":1.0}
                     for p in std["programas"]]
            self._parsed_data = {"programas": progs,
                                 "total_exigido":0, "total_reconocido":0, "total_descuento":0}

        output = self._out_var.get().strip()
        if not output:
            messagebox.showerror("Error", "Especifique la ruta de salida.")
            return
        try:
            info = self._build_info_dict()
            generar_acta_excel(self._parsed_data, info, output)
            self._status_var.set(f"✅ Acta generada exitosamente:\n{output}")
            if messagebox.askyesno("Éxito", f"Acta generada:\n{output}\n\n¿Abrir el archivo?"):
                import subprocess
                subprocess.run(["open", output])
        except Exception as e:
            messagebox.showerror("Error al generar acta", str(e))
            self._status_var.set(f"❌ Error: {e}")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
